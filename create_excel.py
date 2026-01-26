#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Variance Reduction Analysis"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14)
label_font = Font(bold=True)
number_format = '0.00'
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = "Variance Reduction Analysis - P-K Formula Model"
ws['A1'].font = title_font
ws.merge_cells('A1:F1')

# Input Parameters Section
row = 3
ws[f'A{row}'] = "Input Parameters"
ws[f'A{row}'].font = title_font
row += 1

headers = ["Parameter", "Value", "Unit", "Notes"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1
# Store row numbers for reference
row_clients = row
row_variance = row + 1
row_variance2 = row + 2
row_tpt = row + 3
row_service_rate = row + 4
row_arrival_rate = row + 5
row_utilization = row + 6

params = [
    ["Initial Clients", 15, "clients", "Baseline capacity"],
    ["Initial Variance (σ)", 12, "hours", "Standard deviation of service time"],
    ["Initial Variance² (σ²)", 144, "hours²", "Variance of service time"],
    ["Initial TPT", 22, "hours", "Average time per task"],
    ["Service Rate (μ)", f"=1/B{row_tpt}", "clients/hour", "1/TPT"],
    ["Arrival Rate (λ)", f"=B{row_clients}/B{row_tpt}", "clients/hour", "15 clients / 22h"],
    ["Initial Utilization (ρ)", f"=B{row_arrival_rate}/B{row_service_rate}", "ratio", "λ/μ"],
]

for param in params:
    for col, value in enumerate(param, start=1):
        cell = ws.cell(row=row, column=col)
        if isinstance(value, (int, float)):
            cell.value = value
        elif isinstance(value, str) and value.startswith('='):
            cell.value = value
        else:
            cell.value = value
        cell.border = border
        if col == 2 and isinstance(param[1], (int, float)):
            cell.number_format = number_format
        elif col == 2 and isinstance(param[1], str) and param[1].startswith('='):
            cell.number_format = number_format
    row += 1

# P-K Formula Section
row += 1
ws[f'A{row}'] = "P-K Formula: W = (1/μ) + (λ(σ² + 1/μ²)) / (2(1-ρ))"
ws[f'A{row}'].font = Font(italic=True)
ws.merge_cells(f'A{row}:F{row}')

# Variance Reduction Analysis Table
row += 2
ws[f'A{row}'] = "Variance Reduction Analysis"
ws[f'A{row}'].font = title_font
row += 1

headers = ["Variance Reduction %", "Variance (σ)", "Variance² (σ²)", "TPT (W)", "Capacity (Clients)", "ROI Multiplier"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1
variance_reductions = [0, 10, 20, 30, 40, 50, 60, 70, 75, 80]
start_row = row

for vr in variance_reductions:
    # Variance Reduction % - store as number for calculations
    cell = ws.cell(row=row, column=1)
    cell.value = vr
    cell.number_format = '0"%"'
    cell.border = border
    
    # Variance (σ) - Formula: =$B$6*(1-A{row}/100) where B6 is initial variance
    cell = ws.cell(row=row, column=2)
    cell.value = f"=$B${row_variance}*(1-A{row}/100)"
    cell.number_format = number_format
    cell.border = border
    
    # Variance² (σ²) - Formula: =B{row}^2
    cell = ws.cell(row=row, column=3)
    cell.value = f"=B{row}^2"
    cell.number_format = number_format
    cell.border = border
    
    # TPT (W) - P-K formula: =$B$9+($B$10*(C{row}+1/$B$9^2))/(2*(1-$B$11))
    # Using: W = (1/μ) + (λ(σ² + 1/μ²)) / (2(1-ρ))
    cell = ws.cell(row=row, column=4)
    cell.value = f"=$B${row_service_rate}+($B${row_arrival_rate}*(C{row}+1/$B${row_service_rate}^2))/(2*(1-$B${row_utilization}))"
    cell.number_format = number_format
    cell.border = border
    
    # Capacity - Hockey stick curve: =15+0.00306*(A{row})^2
    cell = ws.cell(row=row, column=5)
    cell.value = f"=15+0.00306*(A{row})^2"
    cell.number_format = number_format
    cell.border = border
    
    # ROI Multiplier - Formula: =E{row}/$B$5
    cell = ws.cell(row=row, column=6)
    cell.value = f"=E{row}/$B${row_clients}"
    cell.number_format = number_format
    cell.border = border
    
    row += 1

# Find the row with 70% reduction (7th row after start_row, since 0% is first)
row_70 = start_row + 7  # 0, 10, 20, 30, 40, 50, 60, 70 (index 7)

# Key Metrics at 70% Section
row += 1
ws[f'A{row}'] = "Key Metrics at 70% Variance Reduction"
ws[f'A{row}'].font = title_font
ws.merge_cells(f'A{row}:E{row}')
row += 1

headers = ["Metric", "Before", "After", "Change", "Change %"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1
# Variance (σ)
ws[f'A{row}'] = "Variance (σ)"
ws[f'B{row}'] = f"=$B${row_variance}"
ws[f'C{row}'] = f"=B{row_70}"
ws[f'D{row}'] = f"=C{row}-B{row}"
ws[f'E{row}'] = f"=(C{row}-B{row})/B{row}*100"
for col in range(1, 6):
    ws.cell(row=row, column=col).border = border
    if col > 1:
        ws.cell(row=row, column=col).number_format = number_format
row += 1

# Variance² (σ²)
ws[f'A{row}'] = "Variance² (σ²)"
ws[f'B{row}'] = f"=$B${row_variance2}"
ws[f'C{row}'] = f"=C{row_70}"
ws[f'D{row}'] = f"=C{row}-B{row}"
ws[f'E{row}'] = f"=(C{row}-B{row})/B{row}*100"
for col in range(1, 6):
    ws.cell(row=row, column=col).border = border
    if col > 1:
        ws.cell(row=row, column=col).number_format = number_format
row += 1

# TPT (W)
ws[f'A{row}'] = "TPT (W)"
ws[f'B{row}'] = f"=$B${row_tpt}"
ws[f'C{row}'] = f"=D{row_70}"
ws[f'D{row}'] = f"=C{row}-B{row}"
ws[f'E{row}'] = f"=(C{row}-B{row})/B{row}*100"
for col in range(1, 6):
    ws.cell(row=row, column=col).border = border
    if col > 1:
        ws.cell(row=row, column=col).number_format = number_format
row += 1

# Capacity
ws[f'A{row}'] = "Capacity"
ws[f'B{row}'] = f"=$B${row_clients}"
ws[f'C{row}'] = f"=E{row_70}"
ws[f'D{row}'] = f"=C{row}-B{row}"
ws[f'E{row}'] = f"=(C{row}-B{row})/B{row}*100"
for col in range(1, 6):
    ws.cell(row=row, column=col).border = border
    if col > 1:
        ws.cell(row=row, column=col).number_format = number_format
row += 1

# ROI Multiplier
ws[f'A{row}'] = "ROI Multiplier"
ws[f'B{row}'] = "=1"
ws[f'C{row}'] = f"=F{row_70}"
ws[f'D{row}'] = f"=C{row}-B{row}"
ws[f'E{row}'] = f"=(C{row}-B{row})/B{row}*100"
for col in range(1, 6):
    ws.cell(row=row, column=col).border = border
    if col > 1:
        ws.cell(row=row, column=col).number_format = number_format

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 15

# Save the file
wb.save('variance_reduction_analysis.xlsx')
print("Excel file created successfully: variance_reduction_analysis.xlsx")
